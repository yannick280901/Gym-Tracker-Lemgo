from flask import Flask, render_template, request, redirect, url_for, session
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__)
app.secret_key = "super-geheim-123"  # kannst du ändern

USERNAME = "yannick.2809"
PASSWORD = "tbv1911"

EXCEL_DATEI = "daten.xlsx"

def init_excel():
    if not os.path.exists(EXCEL_DATEI):
        wb = Workbook()
        ws = wb.active
        ws.append(["Datum", "Übung", "Gewicht (kg)", "Wiederholungen"])
        wb.save(EXCEL_DATEI)

@app.route("/", methods=["GET", "POST"])
def index():
    if not session.get("angemeldet"):
        return redirect(url_for("login"))

    init_excel()
    saved = False

    if request.method == "POST":
        uebung = request.form["uebung"]
        gewicht = request.form["gewicht"]
        wdh = request.form["wdh"]
        datum = datetime.now().strftime("%Y-%m-%d %H:%M")

        wb = load_workbook(EXCEL_DATEI)
        ws = wb.active
        ws.append([datum, uebung, gewicht, wdh])
        wb.save(EXCEL_DATEI)

        return redirect(url_for("index", saved="1"))

    uebungen = [
        "Bankdrücken", "Bauch", "Bizeps (EGYM)","Beinbeuger","Hackenschmidt", "Bizeps Hantel",
        "Brustpresse", "Brustpresse (EGYM)", "Brustpresse schräg", "Butterfly",
        "Reverse Butterfly", "Lastzug (EGYM)", "Rudern", "Rudern EGYM",
        "Schulterdrücken (EGYM)", "Trizeps", "Trizeps EGYM", "Unterer Rücken"
    ]
    saved = request.args.get("saved") == "1"
    return render_template("index.html", uebungen=uebungen, saved=saved)

@app.route("/verlauf", methods=["GET"])
def verlauf():
    if not session.get("angemeldet"):
        return redirect(url_for("login"))

    init_excel()
    wb = load_workbook(EXCEL_DATEI)
    ws = wb.active

    daten = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        daten.append({
            "datum": row[0],
            "uebung": row[1],
            "gewicht": row[2],
            "wdh": row[3]
        })

    daten.reverse()

    gefiltert_uebung = request.args.get("uebung", "Alle")
    von_datum_str = request.args.get("von")
    bis_datum_str = request.args.get("bis")

    gefiltert = []
    for eintrag in daten:
        datum_obj = datetime.strptime(eintrag["datum"], "%Y-%m-%d %H:%M")

        innerhalb_datum = True
        if von_datum_str:
            von = datetime.strptime(von_datum_str, "%Y-%m-%d")
            if datum_obj < von:
                innerhalb_datum = False
        if bis_datum_str:
            bis = datetime.strptime(bis_datum_str, "%Y-%m-%d")
            if datum_obj > bis:
                innerhalb_datum = False

        richtige_uebung = (gefiltert_uebung == "Alle" or eintrag["uebung"] == gefiltert_uebung)

        if innerhalb_datum and richtige_uebung:
            gefiltert.append(eintrag)

    uebungen = sorted(list(set(e["uebung"] for e in daten)))
    uebungen.insert(0, "Alle")

    return render_template("verlauf.html",
                           daten=gefiltert,
                           uebungen=uebungen,
                           ausgewaehlt=gefiltert_uebung,
                           von=von_datum_str or "",
                           bis=bis_datum_str or "")

@app.route("/loeschen/<int:index>", methods=["POST"])
def loeschen(index):
    if not session.get("angemeldet"):
        return redirect(url_for("login"))

    init_excel()
    wb = load_workbook(EXCEL_DATEI)
    ws = wb.active

    row_to_delete = len(ws["A"]) - index
    if row_to_delete >= 2:
        ws.delete_rows(row_to_delete)
        wb.save(EXCEL_DATEI)

    return redirect(url_for("verlauf"))

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        if request.form["username"] == USERNAME and request.form["password"] == PASSWORD:
            session["angemeldet"] = True
            return redirect(url_for("index"))
        else:
            return "<h3>❌ Falsche Zugangsdaten</h3><a href='/login'>Zurück</a>"
    return '''
    <h2>🔐 Login</h2>
    <form method="post">
        Benutzername: <input type="text" name="username"><br>
        Passwort: <input type="password" name="password"><br><br>
        <input type="submit" value="Einloggen">
    </form>
    '''

@app.route("/logout")
def logout():
    session.pop("angemeldet", None)
    return redirect(url_for("login"))

app.run(host="0.0.0.0", port=81)
